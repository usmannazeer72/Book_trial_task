"""
Ingestion Service - Load book data from Google Sheets or Excel
"""

import openpyxl
from datetime import datetime
import logging
from typing import List, Dict
from src.database.schema import DatabaseManager, create_book_record
from src.config.settings import config

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class IngestionService:
    """Service to ingest book data from various sources"""
    
    def __init__(self):
        self.db_manager = DatabaseManager()
        self.db = self.db_manager.connect()
    
    def ingest_from_excel(self, file_path: str) -> List[str]:
        """
        Ingest books from local Excel file
        
        Expected columns:
        - title (required)
        - notes_on_outline_before (required)
        - notes_on_outline_after (optional)
        - status_outline_notes (optional)
        
        Returns list of book_ids created
        """
        logger.info(f"Reading Excel file: {file_path}")
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Get header row to map column names to indices
            headers = [cell.value for cell in ws[1]]
            
            # Validate required columns
            required_columns = ['title', 'notes_on_outline_before']
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                raise ValueError(f"Missing required columns: {missing_columns}")
            
            # Get column indices
            title_idx = headers.index('title')
            notes_before_idx = headers.index('notes_on_outline_before')
            
            book_ids = []
            
            # Iterate through rows (skip header)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                title = str(row[title_idx]).strip() if row[title_idx] else ""
                notes_before = str(row[notes_before_idx]).strip() if row[notes_before_idx] else ""
                
                # Skip if title is empty
                if not title or title.lower() == 'none':
                    logger.warning(f"Skipping row {row_idx}: empty title")
                    continue
                
                # Check if book already exists
                existing_book = self.db.books.find_one({'title': title})
                
                if existing_book:
                    logger.info(f"Book '{title}' already exists, updating...")
                    book_id = existing_book['book_id']
                    
                    # Update existing record
                    self.db.books.update_one(
                        {'book_id': book_id},
                        {
                            '$set': {
                                'notes_on_outline_before': notes_before,
                                'updated_at': datetime.utcnow()
                            }
                        }
                    )
                else:
                    # Create new book record
                    book_record = create_book_record(title, notes_before)
                    book_id = book_record['book_id']
                    
                    self.db.books.insert_one(book_record)
                    logger.info(f"✓ Created book: '{title}' (ID: {book_id})")
                
                book_ids.append(book_id)
            
            logger.info(f"✓ Ingested {len(book_ids)} books from Excel")
            return book_ids
        
        except Exception as e:
            logger.error(f"Error ingesting from Excel: {str(e)}")
            raise
    
    def ingest_from_google_sheets(self, spreadsheet_id: str = None, worksheet_name: str = "Books") -> List[str]:
        raise NotImplementedError("Google Sheets ingestion has been removed. Use Excel ingestion instead.")
    
    def get_pending_books(self) -> List[Dict]:
        """Get books that are pending outline generation"""
        books = list(self.db.books.find({
            'notes_on_outline_before': {'$ne': ''},
            'status_outline_notes': {'$in': ['', 'pending']}
        }))
        
        logger.info(f"Found {len(books)} books pending outline generation")
        return books
    
    def close(self):
        """Close database connection"""
        self.db_manager.close()


if __name__ == "__main__":
    # Test ingestion
    service = IngestionService()
    
    # Example: Ingest from Excel
    try:
        book_ids = service.ingest_from_excel(config.app.input_excel_path)
        print(f"Successfully ingested {len(book_ids)} books")
    except Exception as e:
        print(f"Error: {e}")
    
    service.close()
